import openpyxl as oxl
import os
import pandas as pd


AM_list = ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ' ]
PM_list = ['AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ' ]


file = 'Site2_Kanyamazane_Toad street_2019.xlsx'
# Import excel to openpyxl
wb_oxl = oxl.load_workbook(file)


# Import excel file to pandas and clean
# Extract the data from the source file into python
def excel_pd_clean(file, sheet):
    global ts_label
    global q_label
    wb_pd = pd.read_excel(file, sheet_name=sheet)
    if 'stamp' in wb_pd.columns[0].lower():
        ts_label = wb_pd.columns[0]                                         # Specify the label from the document, in case its misspelled
        q_label = wb_pd.columns[1]                                          # Specify the label from the document, in case its misspelled

    wb_pd[ts_label] = wb_pd[ts_label].apply(lambda x: pd.to_datetime(x))    # Convert to pandas timestamp
    return wb_pd


# Add the AM and PM thingy
def add_AM_PM(wb_pd, wb_oxl, sheet):
    AM_PM_list = list(wb_pd[ts_label].dt.strftime('%p').values)

    for row in wb_oxl[sheet].iter_rows(min_row=wb_oxl[sheet].min_row, max_row=wb_oxl[sheet].max_row, min_col=1, max_col=2):
        for cell in row:
            if cell.column_letter == 'A':
                try:
                    AM_PM_val = cell.value.strftime('%p')
                    wb_oxl[sheet]['C' + str(cell.row)] = AM_PM_val
                except Exception:
                    pass


# Generate a list of days in the month
def days_in_sht(wb_pd):
    days_in_sheet = []
    for i, row in wb_pd[ts_label].iteritems():
        if row.day not in days_in_sheet:
            days_in_sheet.append(row.day)
    days_in_sheet = list(range(min(days_in_sheet), max(days_in_sheet)+1)) # Redo this part to make sure their are no missing days before and after
    return days_in_sheet


def add_data_to_new_excel (wb_pd, days_in_sheet):
    month = wb_pd[ts_label].iloc[10].month          # It's lazy, but I just sample from the tenth row and grab the month number
    year = wb_pd[ts_label].iloc[10].year            # Lazy, grabbed the year

    AM = (wb_pd
          .set_index(ts_label)
          .between_time('0:00', '12:00')
          .reset_index()
          .copy()
         )
    PM = (wb_pd
          .set_index(ts_label)
          .between_time('12:00', '0:00')
          .reset_index()
          .copy()
          )


    for day in days_in_sheet:
#         print (day)
        section = list(AM[(AM[ts_label] > pd.Timestamp(year=year, month=month, day=day)) & (AM[ts_label] < pd.Timestamp(year=year, month=month, day=day) + pd.DateOffset(1))][q_label].values)
        col = AM_list[day-1]
        wb_oxl[sheet][col + '1'] = pd.Timestamp(year=year, month=month, day=day).strftime('%Y-%b-%d')
        wb_oxl[sheet][col + '2'] = 'Average Q (L/s)'

        for index, val_q in enumerate(section):
            wb_oxl[sheet][col + str(3+index)] = float(val_q)


    for day in days_in_sheet:
        section = list(PM[(PM[ts_label] > pd.Timestamp(year=year, month=month, day=day)) & (PM[ts_label] < pd.Timestamp(year=year, month=month, day=day) + pd.DateOffset(1))][q_label].values)
        col = PM_list[day-1]
        wb_oxl[sheet][col + '1'] = pd.Timestamp(year=year, month=month, day=day).strftime('%Y-%b-%d')
        wb_oxl[sheet][col + '2'] = 'Average Q (L/s)'

        for index, val_q in enumerate(section):
            wb_oxl[sheet][col + str(3+index)] = float(val_q)


sheet_list = wb_oxl.sheetnames
for sheet in sheet_list:
    wb_pd = excel_pd_clean(file, sheet)      # Pull excel file into pandas dataframe and clean timestamps
    add_AM_PM(wb_pd, wb_oxl, sheet)          # Add AM and PM to olx doc (still to be saved)
    wb_pd = wb_pd[wb_pd[q_label] > 0.01 ]    # Remove all rows with zero
    days_in_sheet = days_in_sht(wb_pd)       #create a list of days to find right column letter and parse correctly
    add_data_to_new_excel (wb_pd, days_in_sheet)


wb_oxl.save('Data_Analysis_' + file[:-5]  + '.xlsx')